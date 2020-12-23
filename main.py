"""
Main file of calendar-sync. Originally written by
Julian van Doorn <jvdoorn@antarc.com>.
"""

import os

from googleapiclient.discovery import build as get_service
from openpyxl import load_workbook
from openpyxl.cell import Cell

from appointment import Appointment, get_appointment_type, get_begin_time, get_end_time
from config import CALENDAR_ID, FIRST_COLUMN, FIRST_ROW, SCHEDULE, STORAGE_FILE
from utils import get_credentials, get_next_cell


def get_calendar_service():
    credentials = get_credentials()
    return get_service('calendar', 'v3', credentials=credentials)


def load_cached_appointments_from_disk():
    stored_appointments = {}

    if os.path.exists(STORAGE_FILE):
        with open(STORAGE_FILE) as f:
            for line in f:
                (checksum, uid) = line.split()
                stored_appointments[checksum] = uid

    return stored_appointments


def load_workbook_from_disk():
    return load_workbook(SCHEDULE).active


def get_first_cell(workbook):
    return Cell(workbook, row=FIRST_ROW, column=FIRST_COLUMN)


def get_appointment_titles_from_cell(cell):
    content = workbook[cell.coordinate].value
    return [] if content is None else content.split(' / ')


if __name__ == '__main__':
    calendar = get_calendar_service()
    workbook = load_workbook_from_disk()

    calendar_appointments = load_cached_appointments_from_disk()
    workbook_appointments = {}

    # These are the appointments that have to be created or looked up in the cache
    create_appointments = []
    # The appointments of the previous iteration
    previous_appointments = []

    # Fetch the first cell
    cell = get_first_cell(workbook)

    while cell:
        appointments = []
        titles = get_appointment_titles_from_cell(cell)
        if len(titles) > 0:
            # Determine the appointment type, begin time and end time.
            appointment_type = get_appointment_type(workbook[cell.coordinate])
            appointment_begin_time = get_begin_time(cell, appointment_type)
            appointment_end_time = get_end_time(cell, appointment_type, workbook)

            for title in titles:
                # Check if we should merge with a previous appointment
                found = False
                for i in range(len(previous_appointments)):
                    previous_appointment = previous_appointments[i]
                    if title == previous_appointment.title:
                        previous_appointment.appointment_end_time = appointment_end_time
                        appointments.append(previous_appointment)
                        previous_appointments.pop(i)
                        found = True
                        break
                # If not, create a new appointment
                if not found:
                    appointments.append(
                        Appointment(title, appointment_type, appointment_begin_time, appointment_end_time))

            # Add any previous appointments (that we did not touch)
            create_appointments += previous_appointments
            # Determine if there is a next cell
            if get_next_cell(cell, workbook) is None:
                # If there is not this was the last iteration and we should add the current appointments
                create_appointments += appointments
            # Update the previous appointments
            previous_appointments = appointments
        else:
            # The cell was empty add the previous appointments and clear the array
            create_appointments += previous_appointments
            previous_appointments = []

        # Get the next cell
        cell = get_next_cell(cell, workbook)

    # Create all the new or updated appointments
    for appointment in create_appointments:
        # Check if we previously created the appointment
        if appointment.checksum() in calendar_appointments:
            # Add it to the new cache
            workbook_appointments[appointment.checksum()] = calendar_appointments[appointment.checksum()]
            # Remove it from the old cache (performance)
            calendar_appointments.pop(appointment.checksum())
        elif appointment.checksum(True) in calendar_appointments:
            # Add it to the new cache
            workbook_appointments[appointment.checksum()] = calendar_appointments[appointment.checksum(True)]
            # Remove it from the old cache (performance)
            calendar_appointments.pop(appointment.checksum(True))
        else:
            # If we did not previously create the appointment we will do so now
            try:
                # Create the event
                event = calendar.events().insert(calendarId=CALENDAR_ID, body=appointment.serialize()).execute(
                    num_retries=10)
                # Add it to the new cache
                workbook_appointments[appointment.checksum()] = event.get('id')
                # Let the user know we created a new event
                print(f'Created event with checksum {appointment.checksum()} and id {event.get("id")}')
            except Exception as e:
                print(f'Error creating event with checksum {appointment.checksum()}: {e}.')

    # Delete any appointments that have been removed or updated in the spreadsheet.
    for uid in calendar_appointments.values():
        try:
            # Delete the old event
            calendar.events().delete(calendarId=CALENDAR_ID, eventId=uid).execute(num_retries=10)
            # Let the user know we deleted an event
            print(f'Deleted event with uid {uid}.')
        except Exception as e:
            print(f'Error deleting event with uid {uid}: {e}.')

    # Save the updated storage to disk
    with open(STORAGE_FILE, 'w') as f:
        for checksum, uid in workbook_appointments.items():
            f.write(f'{checksum} {uid}\n')
