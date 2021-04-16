import os
import pickle
from typing import Dict, List, Tuple

from google.auth.transport.requests import Request
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build as get_service
from openpyxl import load_workbook

from appointment import Appointment
from config import *


def get_credentials():
    credentials = None

    if os.path.exists('token.pickle'):
        with open('token.pickle', 'rb') as token:
            credentials = pickle.load(token)
    if not credentials or not credentials.valid:
        if credentials and credentials.expired and credentials.refresh_token:
            credentials.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                'credentials.json', SCOPES)
            credentials = flow.run_local_server(port=0)
        with open('token.pickle', 'wb') as token:
            pickle.dump(credentials, token)

    return credentials


def get_calendar_service():
    credentials = get_credentials()
    return get_service('calendar', 'v3', credentials=credentials)


def get_stored_appointments() -> Dict[str, Tuple[str, bool]]:
    stored_appointments = {}

    if not os.path.exists(STORAGE_FILE):
        return stored_appointments

    with open(STORAGE_FILE) as storage:
        for line in storage:
            try:
                (checksum, event_id, date) = line.split()
            except ValueError:
                checksum, event_id, date = *line.split(), None

            try:
                historic = datetime.datetime.strptime(date, DATE_FORMAT) < datetime.datetime.now()
            except (AttributeError, TypeError):
                historic = False

            stored_appointments[checksum] = (event_id, historic)
    return stored_appointments


def save_appointments_to_cache(appointments: List[Appointment]):
    with open(STORAGE_FILE, 'w') as f:
        for appointment in appointments:
            f.write(
                f'{appointment.checksum} {appointment.remote_event_id} {appointment.end_time.strftime(DATE_FORMAT)}\n')


def delete_remote_appointments(event_ids: list, calendar):
    for event_id in event_ids:
        delete_appointment(calendar, event_id)


def load_workbook_from_disk():
    return load_workbook(SCHEDULE).active


def update_end_time(previous_appointment, new_end_time):
    previous_appointment.end_time = new_end_time


def create_appointment(calendar, appointment):
    try:
        event = calendar.events().insert(calendarId=CALENDAR_ID, body=appointment.serialize()).execute(num_retries=10)
        print(f"Created a new event {appointment}.")
        return event.get('id')
    except Exception as e:
        print(f'Error creating event {appointment}: {e}.')
        return None


def delete_appointment(calendar, event_id):
    try:
        calendar.events().delete(calendarId=CALENDAR_ID, eventId=event_id).execute(num_retries=10)
        print(f'Deleted event with event ID {event_id}.')
    except Exception as e:
        print(f'Error deleting event with event ID {event_id}: {e}.')
