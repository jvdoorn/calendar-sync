import datetime
import logging
from typing import List, Union

from openpyxl import load_workbook
from openpyxl.cell import Cell, MergedCell
from openpyxl.styles.colors import Color
from openpyxl.worksheet.worksheet import Worksheet

from appointment import Appointment, AppointmentMeta, AppointmentType
from config import BEGIN_TIMES, END_TIMES
from constants import FIRST_DATE_CELL, FIRST_SCHEDULE_COLUMN, FIRST_SCHEDULE_ROW, LAST_SCHEDULE_COLUMN


class ScheduleCell:
    def __init__(self, parent_cell: Cell, reference_date):
        self._parent_cell: Cell = parent_cell
        self._reference_date = reference_date

        self.last_column: int = parent_cell.column
        self.last_row: int = parent_cell.row

    @property
    def first_column(self) -> int:
        return self._parent_cell.column

    @property
    def first_row(self) -> int:
        return self._parent_cell.row

    @property
    def type(self) -> AppointmentType:
        theme = self.color.theme
        rgb = str(self.color.rgb)

        if rgb == 'FF5B9BD5' or theme == 8:
            return AppointmentType.CAMPUS
        elif self.color.theme == 5:
            return AppointmentType.EXAM
        elif rgb == '00000000':
            return AppointmentType.CAMPUS
        elif rgb == 'FFFFC000' or theme == 7:
            return AppointmentType.HOLIDAY
        else:
            logging.warning(f'WARNING: failed to determine appointment type for {self.value} with color {self.color}.')
            return AppointmentType.EMPTY

    @property
    def color(self) -> Color:
        return self._parent_cell.fill.fgColor

    @property
    def value(self):
        return self._parent_cell.value

    @property
    def titles(self):
        return [] if self.value is None else self.value.split(' / ')

    def _get_time(self, last: bool = False):
        row = self.first_row if not last else self.last_row
        column = self.first_column if not last else self.last_column

        time_index = (column - FIRST_SCHEDULE_COLUMN) % 9
        days_delta = (row - FIRST_SCHEDULE_ROW) * 7 + (column - FIRST_SCHEDULE_COLUMN) // 9

        date = self._reference_date + datetime.timedelta(days=days_delta)
        time = BEGIN_TIMES[time_index] if not last else END_TIMES[time_index]

        return date.replace(hour=time[0], minute=time[1])

    @property
    def begin_time(self) -> datetime.datetime:
        return self._get_time()

    @property
    def end_time(self) -> datetime.datetime:
        return self._get_time(True)


class Schedule:
    def __init__(self, file: str):
        self._worksheets: List[Worksheet] = load_workbook(file).worksheets

    def __iter__(self):
        self._cells: List[Cell] = []
        return self

    def __next__(self) -> ScheduleCell:
        if len(self._cells) == 0:
            if len(self._worksheets) == 0:
                raise StopIteration

            self._worksheet = self._worksheets.pop()
            logging.info(f'Processing worksheet {self._worksheet.title}.')

            rows = self._worksheet.iter_rows(min_row=FIRST_SCHEDULE_ROW, min_col=FIRST_SCHEDULE_COLUMN,
                                             max_col=LAST_SCHEDULE_COLUMN)
            self._cells = [cell for row in rows for cell in row]

            date_cell: Cell = self._worksheet[FIRST_DATE_CELL]
            self._reference_date = date_cell.value

        current_cell = ScheduleCell(self._cells.pop(0), self._reference_date)
        while len(self._cells) > 0 and isinstance(self._cells[0], MergedCell):
            cell = self._cells.pop(0)

            current_cell.last_column = cell.column
            current_cell.last_row = cell.row

        return current_cell

    def get_appointments_from_workbook(self, appointment_meta) -> List[Appointment]:
        appointments_in_workbook: List[Appointment] = []

        previous_appointments: List[Appointment] = []
        for cell in iter(self):
            current_appointments: List[Appointment] = []

            for title in cell.titles:
                title = title.strip()
                appointment: Union[Appointment, None] = None

                for previous_appointment in previous_appointments:
                    if previous_appointment.title == title and previous_appointment.type == cell.type:
                        appointment = previous_appointment
                        appointment.end_time = cell.end_time

                        previous_appointments.remove(previous_appointment)
                        break

                if appointment is None:
                    try:
                        meta = appointment_meta[title]
                    except KeyError:
                        meta = AppointmentMeta(title=title)
                        appointment_meta[title] = meta

                    appointment = Appointment(meta, cell.type, cell.begin_time, cell.end_time)

                current_appointments.append(appointment)

            appointments_in_workbook += previous_appointments
            previous_appointments = current_appointments

        appointments_in_workbook += previous_appointments
        return appointments_in_workbook
